import sqlite3
import openpyxl
import logging
import re
from datetime import datetime

# Настройка логирования
logger = logging.getLogger(__name__)

def get_connection():
    return sqlite3.connect('donor_bot.db', timeout=10)

def init_db():
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            # Таблица users: fio вместо name и surname
            cursor.execute('''CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER UNIQUE,
                phone TEXT UNIQUE,
                fio TEXT,
                category TEXT,
                user_group TEXT,
                social_contacts TEXT,
                dkm BOOLEAN DEFAULT 0,
                consent BOOLEAN DEFAULT 0,
                profile_status TEXT DEFAULT 'pending'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                time TEXT,
                location TEXT,
                description TEXT,
                capacity INTEGER,
                status TEXT DEFAULT 'active'
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS registrations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                event_id INTEGER,
                status TEXT DEFAULT 'registered',
                attended BOOLEAN DEFAULT 0,
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (event_id) REFERENCES events(id)
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS reminders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                event_id INTEGER,
                reminder_date TEXT,
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (event_id) REFERENCES events(id)
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS donations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                date TEXT,
                center TEXT,
                FOREIGN KEY (user_id) REFERENCES users(id)
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS non_attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                registration_id INTEGER,
                reason TEXT,
                FOREIGN KEY (registration_id) REFERENCES registrations(id)
            )''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS questions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                text TEXT,
                answered BOOLEAN DEFAULT 0,
                timestamp TEXT,
                FOREIGN KEY (user_id) REFERENCES users(id)
            )''')
            # Таблица для админов
            cursor.execute('''CREATE TABLE IF NOT EXISTS admins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER UNIQUE
            )''')
            # Начальное заполнение админов, если таблица пуста
            cursor.execute('SELECT COUNT(*) FROM admins')
            if cursor.fetchone()[0] == 0:
                initial_admins = [123456789, 1653833795, 1191457973]
                for admin_id in initial_admins:
                    cursor.execute('INSERT OR IGNORE INTO admins (telegram_id) VALUES (?)', (admin_id,))
            conn.commit()
            logger.info("База данных успешно инициализирована")
    except sqlite3.Error as e:
        logger.error(f"Ошибка инициализации базы данных: {e}")
        raise

def import_from_excel():
    try:
        logger.info("Начало импорта данных из Excel")
        wb = openpyxl.load_workbook('База ДД.xlsx')
        sheet = wb.active
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM users')
            if cursor.fetchone()[0] > 0:
                logger.info("Таблица users не пуста, пропуск импорта")
                return
            imported_count = 0
            skipped_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                fio = str(row[0]).strip() if row[0] else ''
                if not fio:
                    logger.warning(f"Пропущена строка {row}: пустое ФИО")
                    skipped_count += 1
                    continue
                user_group = str(row[1]).strip() if row[1] else ''
                category = ('сотрудник' if 'сотрудник' in user_group.lower() or 'инженер' in user_group.lower()
                            else 'студент' if re.match(r'^[А-Я]\d{2}-\d{3}$', user_group)
                            else 'внешний')
                social_contacts = str(row[7]).strip() if row[7] else None
                phone = str(row[8]).strip() if row[8] else None
                if not phone:
                    logger.warning(f"Пропущена строка для ФИО {fio}: отсутствует номер телефона")
                    skipped_count += 1
                    continue
                cursor.execute('SELECT id FROM users WHERE phone = ?', (phone,))
                if cursor.fetchone():
                    logger.warning(f"Пропущена строка для ФИО {fio}: телефон {phone} уже существует")
                    skipped_count += 1
                    continue
                try:
                    cursor.execute('''INSERT OR REPLACE INTO users 
                        (phone, fio, category, user_group, social_contacts, profile_status)
                        VALUES (?, ?, ?, ?, ?, 'approved')''',
                        (phone, fio, category, user_group, social_contacts))
                    user_id = cursor.lastrowid
                    count_gavrilov = int(row[2]) if row[2] else 0
                    count_fmba = int(row[3]) if row[3] else 0
                    last_gavrilov = row[5] if row[5] else None
                    last_fmba = row[6] if row[6] else None
                    for _ in range(count_gavrilov):
                        cursor.execute('INSERT INTO donations (user_id, date, center) VALUES (?, ?, ?)',
                                       (user_id, last_gavrilov or 'unknown', 'Гаврилова'))
                    for _ in range(count_fmba):
                        cursor.execute('INSERT INTO donations (user_id, date, center) VALUES (?, ?, ?)',
                                       (user_id, last_fmba or 'unknown', 'ФМБА'))
                    imported_count += 1
                    logger.info(f"Импортирован пользователь: {fio}, телефон: {phone}, категория: {category}, группа: {user_group}")
                except sqlite3.Error as e:
                    logger.error(f"Ошибка при вставке пользователя {fio} (телефон: {phone}): {e}")
                    skipped_count += 1
            conn.commit()
            logger.info(f"Импорт завершен: импортировано {imported_count} записей, пропущено {skipped_count} записей")
    except FileNotFoundError:
        logger.error("Файл 'База ДД.xlsx' не найден")
        raise
    except Exception as e:
        logger.error(f"Ошибка импорта Excel: {e}")
        raise

def get_user_by_phone(phone):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE phone = ?', (phone,))
        return cursor.fetchone()

def get_consent_by_phone(phone):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT consent FROM users WHERE phone = ?', (phone,))
        result = cursor.fetchone()
        return result[0] if result else None

def update_consent_by_phone(phone, consent):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE users SET consent = ? WHERE phone = ?', (consent, phone))
        conn.commit()

def save_or_update_user(telegram_id, phone, fio, category, user_group, social_contacts):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM users WHERE phone = ?', (phone,))
        existing = cursor.fetchone()
        if existing:
            cursor.execute('''UPDATE users SET 
                telegram_id = ?, fio = ?, category = ?, user_group = ?, social_contacts = ?, profile_status = 'pending'
                WHERE phone = ?''',
                (telegram_id, fio, category, user_group, social_contacts, phone))
            logger.info(f"Обновлен профиль для телефона {phone}, отправлен на модерацию")
        else:
            cursor.execute('''INSERT INTO users 
                (telegram_id, phone, fio, category, user_group, social_contacts, profile_status)
                VALUES (?, ?, ?, ?, ?, ?, 'pending')''',
                (telegram_id, phone, fio, category, user_group, social_contacts))
            logger.info(f"Создан новый профиль для {fio}, отправлен на модерацию")
        conn.commit()

def get_profile_status_by_telegram_id(telegram_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT profile_status FROM users WHERE telegram_id = ?', (telegram_id,))
        result = cursor.fetchone()
        return result[0] if result else None

def get_active_events():
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id, date, time, location, description, capacity FROM events WHERE status = "active"')
        return cursor.fetchall()

def get_user_by_telegram_id(telegram_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            'SELECT id, fio, category, user_group, social_contacts, dkm, profile_status FROM users WHERE telegram_id = ?',
            (telegram_id,))
        return cursor.fetchone()

def get_donations_count_by_center(user_id, center):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM donations WHERE user_id = ? AND center = ?', (user_id, center))
        return cursor.fetchone()[0]

def get_last_donation(user_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT MAX(date), center FROM donations WHERE user_id = ? GROUP BY center ORDER BY date DESC LIMIT 1', (user_id,))
        return cursor.fetchone()

def get_donations_history(user_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT date, center FROM donations WHERE user_id = ? ORDER BY date DESC', (user_id,))
        return cursor.fetchall()

def get_user_registrations(user_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT e.id, e.date, e.time, e.description FROM registrations r JOIN events e ON r.event_id = e.id WHERE r.user_id = ? AND r.status = "registered"', (user_id,))
        return cursor.fetchall()

def get_user_registrations_count(user_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM registrations WHERE user_id = ? AND status = "registered"', (user_id,))
        return cursor.fetchone()[0]

def get_admin_stats():
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM users WHERE profile_status = "approved"')
        users_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM events')
        events_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM registrations WHERE status = "registered"')
        reg_count = cursor.fetchone()[0]
        return users_count, events_count, reg_count

def get_pending_users():
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id, fio, user_group, social_contacts FROM users WHERE profile_status = "pending"')
        return cursor.fetchall()

def update_profile_status(user_id, status):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE users SET profile_status = ? WHERE id = ?', (status, user_id))
        conn.commit()

def get_telegram_id_by_user_id(user_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT telegram_id FROM users WHERE id = ?', (user_id,))
        result = cursor.fetchone()
        return result[0] if result else None

def add_event(date, time, location, description, capacity):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('INSERT INTO events (date, time, location, description, capacity) VALUES (?, ?, ?, ?, ?)',
                       (date, time, location, description, capacity))
        conn.commit()
        return cursor.lastrowid

def get_consented_users_telegram_ids():
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT telegram_id FROM users WHERE consent = 1')
        return [row[0] for row in cursor.fetchall()]

def get_all_events():
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id, date, time, description, capacity, status FROM events')
        return cursor.fetchall()

def get_registrations_count(event_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM registrations WHERE event_id = ? AND status = "registered"', (event_id,))
        return cursor.fetchone()[0]

def get_attended_count(event_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM registrations WHERE event_id = ? AND attended = 1', (event_id,))
        return cursor.fetchone()[0]

def get_event_status(event_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT status FROM events WHERE id = ?', (event_id,))
        result = cursor.fetchone()
        return result[0] if result else None

def update_event_status(event_id, status):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE events SET status = ? WHERE id = ?', (status, event_id))
        conn.commit()

def get_user_by_id(user_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
        return cursor.fetchone()

def get_users_paginated(limit, offset):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id, fio, user_group FROM users LIMIT ? OFFSET ?', (limit, offset))
        return cursor.fetchall()

def delete_user_by_id(user_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()

def get_all_users_for_export():
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users')
        return cursor.fetchall()

def get_reminders_to_send(current_date):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id, user_id, event_id, reminder_date FROM reminders WHERE reminder_date <= ?',
                       (current_date,))
        return cursor.fetchall()

def get_event_by_id(event_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT date, time, location FROM events WHERE id = ?', (event_id,))
        return cursor.fetchone()

def delete_reminder(reminder_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM reminders WHERE id = ?', (reminder_id,))
        conn.commit()

def get_past_events(today):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM events WHERE date < ? AND status = "active"', (today,))
        return cursor.fetchall()

def get_non_attended_registrations(event_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id, user_id FROM registrations WHERE event_id = ? AND status = "registered" AND attended = 0', (event_id,))
        return cursor.fetchall()

def add_non_attendance_reason(reg_id, reason):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('INSERT INTO non_attendance (registration_id, reason) VALUES (?, ?)', (reg_id, reason))
        conn.commit()

def get_user_id_by_telegram_id(telegram_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM users WHERE telegram_id = ?', (telegram_id,))
        result = cursor.fetchone()
        return result[0] if result else None

def get_event_capacity(event_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT capacity FROM events WHERE id = ?', (event_id,))
        result = cursor.fetchone()
        return result[0] if result else None

def get_event_date(event_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT date FROM events WHERE id = ?', (event_id,))
        result = cursor.fetchone()
        return result[0] if result else None

def add_registration(user_id, event_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('INSERT INTO registrations (user_id, event_id) VALUES (?, ?)', (user_id, event_id))
        conn.commit()

def add_reminder(user_id, event_id, reminder_date):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('INSERT INTO reminders (user_id, event_id, reminder_date) VALUES (?, ?, ?)',
                       (user_id, event_id, reminder_date))
        conn.commit()

def cancel_registration(user_id, event_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE registrations SET status = "cancelled" WHERE user_id = ? AND event_id = ?', (user_id, event_id))
        conn.commit()

def add_donation(user_id, date, center):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('INSERT INTO donations (user_id, date, center) VALUES (?, ?, ?)', (user_id, date, center))
        conn.commit()

def update_dkm(user_id, dkm):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE users SET dkm = ? WHERE id = ?', (dkm, user_id))
        conn.commit()

def get_user_by_name_surname(fio):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM users WHERE fio = ?', (fio,))
        return cursor.fetchone()

def add_question(user_id, text):
    with get_connection() as conn:
        cursor = conn.cursor()
        timestamp = datetime.now().isoformat()
        cursor.execute('INSERT INTO questions (user_id, text, timestamp) VALUES (?, ?, ?)',
                       (user_id, text, timestamp))
        conn.commit()
        return cursor.lastrowid

def get_unanswered_questions():
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id, user_id, text, timestamp FROM questions WHERE answered = 0 ORDER BY timestamp DESC')
        return cursor.fetchall()

def mark_question_answered(question_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE questions SET answered = 1 WHERE id = ?', (question_id,))
        conn.commit()

def get_user_telegram_id(user_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT telegram_id FROM users WHERE id = ?', (user_id,))
        result = cursor.fetchone()
        return result[0] if result else None

# Функция добавления админа (исправлена для работоспособности, с обработкой исключений)
def add_admin(telegram_id):
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('INSERT OR IGNORE INTO admins (telegram_id) VALUES (?)', (telegram_id,))
            conn.commit()
        logger.info(f"Добавлен админ с telegram_id {telegram_id}")
    except sqlite3.Error as e:
        logger.error(f"Ошибка при добавлении админа {telegram_id}: {e}")
        raise

# Функции для бэкапа и восстановления пользователей
def export_users_to_excel(filename='users_backup.xlsx'):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['id', 'telegram_id', 'phone', 'fio', 'category', 'user_group', 'social_contacts', 'dkm', 'consent', 'profile_status'])
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM users')
            users = cursor.fetchall()
            for user in users:
                sheet.append(user)
        wb.save(filename)
        logger.info(f"Экспорт пользователей сохранен в {filename}")
        return filename
    except Exception as e:
        logger.error(f"Ошибка экспорта пользователей: {e}")
        raise

def import_users_from_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM users')  # Очистка таблицы перед импортом (опасная операция)
            imported = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 10:
                    continue
                cursor.execute('''INSERT OR REPLACE INTO users 
                    (id, telegram_id, phone, fio, category, user_group, social_contacts, dkm, consent, profile_status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', row)
                imported += 1
            conn.commit()
            logger.info(f"Восстановлено {imported} пользователей из {filename}")
    except Exception as e:
        logger.error(f"Ошибка импорта пользователей из {filename}: {e}")
        raise