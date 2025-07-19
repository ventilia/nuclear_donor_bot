import asyncio
from datetime import datetime

from aiogram import types, Router, Dispatcher
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardMarkup, KeyboardButton

from src.states.states import AddEventStates, BroadcastState, AnswerQuestionState, AddAdminState, RestoreState
from src.database.db import get_admin_stats, get_pending_users, update_profile_status, get_telegram_id_by_user_id, \
    get_connection, get_user_registrations_count, add_donation, update_dkm, get_user_by_fio, \
    get_donations_count_by_center, get_last_donation
from src.database.db import add_event, get_consented_users_telegram_ids, get_all_events, get_registrations_count
from src.database.db import get_attended_count, get_event_status, update_event_status, get_user_by_id, get_users_paginated
from src.database.db import delete_user_by_id, get_all_users_for_export, add_question, get_unanswered_questions
from src.database.db import mark_question_answered, get_user_telegram_id, import_from_excel, export_users_to_excel, import_users_from_excel
from src.utils.keyboards import is_admin
from src.database.db import logger, add_admin
import openpyxl

admin_router = Router()

@admin_router.message(Command(commands=['add_admin']))
async def add_admin_handler(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    await state.set_state(AddAdminState.telegram_id)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите Telegram ID нового админа:", reply_markup=keyboard)

@admin_router.message(AddAdminState.telegram_id)
async def process_add_admin_id(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.clear()
        await message.answer("Добавление админа отменено.", reply_markup=types.ReplyKeyboardRemove())
        return
    try:
        telegram_id = int(message.text.strip())
        await state.update_data(telegram_id=telegram_id)
        await state.set_state(AddAdminState.confirm)
        keyboard = InlineKeyboardBuilder()
        keyboard.button(text="Подтвердить ✅", callback_data="add_admin_confirm")
        keyboard.button(text="Отмена ❌", callback_data="add_admin_cancel")
        await message.answer(f"Подтвердите добавление админа с ID {telegram_id}?", reply_markup=keyboard.as_markup())
    except ValueError:
        await message.answer("Некорректный ID. Введите число. ⚠️")

@admin_router.callback_query(lambda c: c.data in ['add_admin_confirm', 'add_admin_cancel'], AddAdminState.confirm)
async def confirm_add_admin(callback_query: types.CallbackQuery, state: FSMContext):
    if callback_query.data == 'add_admin_cancel':
        await state.clear()
        await callback_query.message.answer("Добавление отменено.")
        await callback_query.answer()
        return
    data = await state.get_data()
    telegram_id = data.get('telegram_id')
    try:
        add_admin(telegram_id)
        await callback_query.message.answer(f"Админ с ID {telegram_id} добавлен. ✅")
        logger.info(f"Админ {callback_query.from_user.id} добавил нового админа {telegram_id}")
    except Exception as e:
        logger.error(f"Ошибка при добавлении админа {telegram_id}: {e}")
        await callback_query.message.answer("Произошла ошибка. ⚠️")
    await state.clear()
    await callback_query.answer()

@admin_router.message(Command(commands=['backup_users']))
async def backup_users_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    try:
        filename = export_users_to_excel()
        await message.answer_document(types.FSInputFile(filename), caption="Бэкап таблицы users. 📂")
        logger.info(f"Админ {message.from_user.id} создал бэкап users")
    except Exception as e:
        logger.error(f"Ошибка бэкапа users: {e}")
        await message.answer("Ошибка при создании бэкапа. ⚠️")

@admin_router.message(Command(commands=['restore_users']))
async def restore_users_handler(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    await state.set_state(RestoreState.file)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Отправьте файл Excel для восстановления users (опасно: перезапишет данные!): 📂", reply_markup=keyboard)

@admin_router.message(lambda m: m.document and m.document.file_name.endswith('.xlsx'), RestoreState.file)
async def process_restore_file(message: types.Message, state: FSMContext):
    # Локальный импорт bot
    from src.bot import bot
    if not message.document:
        await message.answer("Отправьте файл Excel (.xlsx). ⚠️")
        return
    try:
        file_id = message.document.file_id
        file = await bot.get_file(file_id)
        file_path = file.file_path
        temp_filename = "temp_restore.xlsx"
        await bot.download_file(file_path, temp_filename)
        await state.update_data(filename=temp_filename)
        await state.set_state(RestoreState.confirm)
        keyboard = InlineKeyboardBuilder()
        keyboard.button(text="Подтвердить ✅", callback_data="restore_confirm")
        keyboard.button(text="Отмена ❌", callback_data="restore_cancel")
        await message.answer("Подтвердите восстановление (это перезапишет все данные в users!)?", reply_markup=keyboard.as_markup())
    except Exception as e:
        logger.error(f"Ошибка загрузки файла для восстановления: {e}")
        await message.answer("Ошибка при загрузке файла. ⚠️")

@admin_router.callback_query(lambda c: c.data in ['restore_confirm', 'restore_cancel'], RestoreState.confirm)
async def confirm_restore(callback_query: types.CallbackQuery, state: FSMContext):
    if callback_query.data == 'restore_cancel':
        await state.clear()
        await callback_query.message.answer("Восстановление отменено.")
        await callback_query.answer()
        return
    data = await state.get_data()
    filename = data.get('filename')
    try:
        import_users_from_excel(filename)
        await callback_query.message.answer("Восстановление users завершено. ✅")
        logger.info(f"Админ {callback_query.from_user.id} восстановил users из бэкапа")
    except Exception as e:
        logger.error(f"Ошибка восстановления users: {e}")
        await callback_query.message.answer("Ошибка при восстановлении. ⚠️")
    await state.clear()
    await callback_query.answer()

@admin_router.message(Command(commands=['answer']))
async def answer_handler(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    try:
        questions = get_unanswered_questions()
        if not questions:
            await message.answer("Нет неотвеченных вопросов.")
            return
        keyboard = InlineKeyboardBuilder()
        for q in questions:
            user_tg_id = get_user_telegram_id(q[1])
            keyboard.button(text=f"Вопрос {q[0]} от {user_tg_id}", callback_data=f"ans_{q[0]}")
        await message.answer("Выберите вопрос для ответа:", reply_markup=keyboard.as_markup())
        await state.set_state(AnswerQuestionState.select)
    except Exception as e:
        logger.error(f"Ошибка при получении вопросов для ответа: {e}")
        await message.answer("Произошла ошибка.")

@admin_router.callback_query(lambda c: c.data.startswith('ans_'), AnswerQuestionState.select)
async def select_question(callback_query: types.CallbackQuery, state: FSMContext):
    question_id = int(callback_query.data.split('_')[1])
    await state.update_data(question_id=question_id)
    await state.set_state(AnswerQuestionState.response)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await callback_query.message.answer("Введите ответ на вопрос:", reply_markup=keyboard)
    await callback_query.answer()

@admin_router.message(AnswerQuestionState.response)
async def process_answer_text(message: types.Message, state: FSMContext):
    # Локальный импорт bot внутри функции для избежания циклического импорта
    from src.bot import bot
    if message.text == "Назад 🔙":
        await state.set_state(AnswerQuestionState.select)
        await answer_handler(message, state)
        return
    text = message.text.strip()
    if not text:
        await message.answer("Ответ не может быть пустым. Попробуйте снова.")
        return
    data = await state.get_data()
    question_id = data.get('question_id')
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT user_id FROM questions WHERE id = ?', (question_id,))
            user_id = cursor.fetchone()[0]
        user_tg_id = get_user_telegram_id(user_id)
        if user_tg_id:
            await bot.send_message(user_tg_id, f"Ответ на ваш вопрос от организаторов: {text}")
            mark_question_answered(question_id)
            await message.answer("Ответ отправлен пользователю.", reply_markup=types.ReplyKeyboardRemove())
            logger.info(f"Админ {message.from_user.id} ответил на вопрос {question_id}")
        else:
            await message.answer("Не удалось найти Telegram ID пользователя.")
    except Exception as e:
        logger.error(f"Ошибка при отправке ответа на вопрос {question_id}: {e}")
        await message.answer("Произошла ошибка.")
    await state.clear()

@admin_router.message(Command(commands=['broadcast']))
async def broadcast_handler(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    await state.set_state(BroadcastState.text)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите текст для рассылки всем пользователям:", reply_markup=keyboard)

@admin_router.message(BroadcastState.text)
async def process_broadcast_text(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.clear()
        await message.answer("Рассылка отменена.", reply_markup=types.ReplyKeyboardRemove())
        return
    text = message.text.strip()
    if not text:
        await message.answer("Текст не может быть пустым. Попробуйте снова.")
        return
    await state.update_data(text=text)
    await state.set_state(BroadcastState.photo)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Без фото 📄")],
                  [KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Прикрепите фото (если нужно) или нажмите 'Без фото':", reply_markup=keyboard)

@admin_router.message(BroadcastState.photo)
async def process_broadcast_photo(message: types.Message, state: FSMContext):
    # Локальный импорт bot внутри функции для избежания циклического импорта
    from src.bot import bot
    if message.text == "Назад 🔙":
        await state.set_state(BroadcastState.text)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите текст для рассылки:", reply_markup=keyboard)
        return
    if message.text == "Без фото 📄":
        photo = None
    elif message.photo:
        photo = message.photo[-1].file_id
    else:
        await message.answer("Пожалуйста, прикрепите фото или выберите 'Без фото'.")
        return
    await state.update_data(photo=photo)
    await state.set_state(BroadcastState.confirm)
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="Подтвердить ✅", callback_data="broadcast_confirm")
    keyboard.button(text="Отмена ❌", callback_data="broadcast_cancel")
    await message.answer("Подтвердите рассылку?", reply_markup=keyboard.as_markup())

@admin_router.callback_query(lambda c: c.data in ['broadcast_confirm', 'broadcast_cancel'], BroadcastState.confirm)
async def confirm_broadcast(callback_query: types.CallbackQuery, state: FSMContext):
    # Локальный импорт bot внутри функции для избежания циклического импорта
    from src.bot import bot
    if callback_query.data == 'broadcast_cancel':
        await state.clear()
        await callback_query.message.answer("Рассылка отменена.")
        await callback_query.answer()
        return
    data = await state.get_data()
    text = data.get('text')
    photo = data.get('photo')
    try:
        users = get_consented_users_telegram_ids()
        sent_count = 0
        for tg_id in users:
            try:
                if photo:
                    await bot.send_photo(tg_id, photo, caption=text)
                else:
                    await bot.send_message(tg_id, text)
                sent_count += 1
            except Exception as e:
                logger.warning(f"Ошибка отправки рассылки пользователю {tg_id}: {e}")
        await callback_query.message.answer(f"Рассылка завершена. Отправлено {sent_count} пользователям.")
        logger.info(f"Админ {callback_query.from_user.id} отправил рассылку: {text[:50]}...")
    except Exception as e:
        logger.error(f"Ошибка при рассылке: {e}")
        await callback_query.message.answer("Произошла ошибка при рассылке.")
    await state.clear()
    await callback_query.answer()

@admin_router.message(Command(commands=['admin_stats']))
async def admin_stats_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался получить админскую статистику")
        return
    try:
        users_count, events_count, reg_count = get_admin_stats()
        await message.answer(
            f"Статистика: 📊\nПользователей: {users_count} 👥\nМероприятий: {events_count} 📅\nРегистраций: {reg_count} 📝")
        logger.info(f"Админ {message.from_user.id} запросил статистику")
    except Exception as e:
        logger.error(f"Ошибка при получении админской статистики: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@admin_router.message(Command(commands=['admin_reg']))
async def admin_reg_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался управлять заявками")
        return
    try:
        pending_users = get_pending_users()
        if not pending_users:
            await message.answer("Нет заявок. 📭")
            logger.info("Нет заявок на модерацию")
            return
        for user in pending_users:
            keyboard = InlineKeyboardBuilder()
            keyboard.button(text="Принять ✅", callback_data=f"approve_{user[0]}")
            keyboard.button(text="Отклонить ❌", callback_data=f"reject_{user[0]}")
            await message.answer(
                f"Заявка: {user[1]}\nГруппа: {user[2]}\nСоцсети: {user[3]}",
                reply_markup=keyboard.as_markup())
            logger.info(f"Отображена заявка пользователя {user[1]} для админа {message.from_user.id}")
    except Exception as e:
        logger.error(f"Ошибка при получении заявок на модерацию: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@admin_router.callback_query(lambda c: c.data.startswith('approve_') or c.data.startswith('reject_'))
async def process_profile_action(callback_query: types.CallbackQuery):
    # Локальный импорт bot внутри функции для избежания циклического импорта
    from src.bot import bot
    action, user_id_str = callback_query.data.split('_')
    user_id = int(user_id_str)
    try:
        status = 'approved' if action == 'approve' else 'rejected'
        update_profile_status(user_id, status)
        telegram_id = get_telegram_id_by_user_id(user_id)
        if action == 'approve' and telegram_id is not None:
            await bot.send_message(telegram_id, "Ваш профиль был принят администратором. ✅")
            help_text = ("Вот команды пользователя: 📋\n"
                         "/profilReg - Зарегистрировать профиль ✍️\n"
                         "/reg - Записаться на мероприятие 📅\n"
                         "/profil - Посмотреть/изменить профиль 👤\n"
                         "/stats - Моя статистика 📊\n"
                         "/info - Информационные разделы 📖\n"
                         "/ask - Задать вопрос организаторам ❓\n"
                         "/help - Показать этот список ❓")
            await bot.send_message(telegram_id, help_text)
        logger.info(f"Профиль пользователя ID {user_id} {status} админом {callback_query.from_user.id}")
        await callback_query.answer(f"Профиль {'принят ✅' if action == 'approve' else 'отклонен ❌'}.")
    except Exception as e:
        logger.error(f"Ошибка при обработке профиля ID {user_id}: {e}")
        await callback_query.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@admin_router.message(Command(commands=['admin_help']))
async def admin_help_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался получить админскую помощь")
        return
    await message.answer("/admin_stats - Статистика проекта\n"
                         "/admin_reg - Управление заявками\n"
                         "/add_event - Добавить мероприятие\n"
                         "/stats_event - Статистика мероприятий\n"
                         "/see_profile - Просмотр профилей\n"
                         "/see_profile (числовый аргумент) - Просмотр конкретного пользователя по айди\n"
                         "/import_excel - Импорт из Excel\n"
                         "/upload_stats - Загрузить статистику из Excel\n"
                         "/export_stats - Выгрузить статистику в Excel\n"
                         "/answer - Ответить на вопросы пользователей\n"
                         "/broadcast - Рассылка сообщений всем пользователям\n"
                         "/add_admin - Добавить админа\n"
                         "/backup_users - Бэкап пользователей\n"
                         "/restore_users - Восстановление пользователей из бэкапа\n"
                         "/help - Список пользовательских команд")
    logger.info(f"Админ {message.from_user.id} запросил список админских команд")

@admin_router.message(Command(commands=['add_event']))
async def add_event_handler(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался добавить мероприятие")
        return
    await state.set_state(AddEventStates.date)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите дату (YYYY-MM-DD): 📅", reply_markup=keyboard)
    logger.info(f"Админ {message.from_user.id} начал добавление мероприятия")

@admin_router.message(AddEventStates.date)
async def process_event_date(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.clear()
        await message.answer("Добавление отменено.", reply_markup=types.ReplyKeyboardRemove())
        return
    try:
        event_date = datetime.strptime(message.text, '%Y-%m-%d')
        if event_date < datetime.now().replace(hour=0, minute=0, second=0, microsecond=0):
            await message.answer("Дата уже прошла. Введите будущую дату в формате YYYY-MM-DD. ⚠️")
            logger.warning(f"Попытка создать мероприятие с прошедшей датой {message.text} от админа {message.from_user.id}")
            return
    except ValueError:
        await message.answer("Некорректный формат даты. Пожалуйста, введите дату в формате YYYY-MM-DD. ⚠️")
        logger.warning(f"Некорректный формат даты от админа {message.from_user.id}: {message.text}")
        return
    await state.update_data(date=message.text)
    await state.set_state(AddEventStates.time)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите время (HH:MM): ⏰", reply_markup=keyboard)

@admin_router.message(AddEventStates.time)
async def process_event_time(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(AddEventStates.date)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите дату (YYYY-MM-DD): 📅", reply_markup=keyboard)
        return
    await state.update_data(time=message.text)
    await state.set_state(AddEventStates.location)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите место: 📍", reply_markup=keyboard)

@admin_router.message(AddEventStates.location)
async def process_event_location(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(AddEventStates.time)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите время (HH:MM): ⏰", reply_markup=keyboard)
        return
    await state.update_data(location=message.text)
    await state.set_state(AddEventStates.description)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите описание: 📝", reply_markup=keyboard)

@admin_router.message(AddEventStates.description)
async def process_event_description(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(AddEventStates.location)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите место: 📍", reply_markup=keyboard)
        return
    await state.update_data(description=message.text)
    await state.set_state(AddEventStates.capacity)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите вместимость: 👥", reply_markup=keyboard)

@admin_router.message(AddEventStates.capacity)
async def process_event_capacity(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(AddEventStates.description)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите описание: 📝", reply_markup=keyboard)
        return
    try:
        capacity = int(message.text)
        if capacity <= 0:
            raise ValueError("Вместимость должна быть положительным числом")
    except ValueError:
        await message.answer("Вместимость должна быть числом больше 0. ⚠️")
        logger.warning(f"Некорректная вместимость от админа {message.from_user.id}: {message.text}")
        return
    data = await state.get_data()
    try:
        add_event(data['date'], data['time'], data['location'], data['description'], capacity)
        await state.clear()
        await message.answer("Мероприятие добавлено. ✅", reply_markup=types.ReplyKeyboardRemove())
        asyncio.create_task(send_new_event_notification(data['date'], data['time'], data['location'], data['description']))
        logger.info(f"Админ {message.from_user.id} добавил мероприятие: {data['description']}")
    except Exception as e:
        logger.error(f"Ошибка при добавлении мероприятия: {e}")
        await message.answer("Произошла ошибка при добавлении мероприятия. Попробуйте позже. ⚠️")

async def send_new_event_notification(date, time, location, description):
    # Локальный импорт bot внутри функции для избежания циклического импорта
    from src.bot import bot
    try:
        users = get_consented_users_telegram_ids()
        for telegram_id in users:
            try:
                await bot.send_message(telegram_id, f"Новое мероприятие: {description} 📅\nДата: {date} {time} ⏰\nМесто: {location} 📍\nЗарегистрируйтесь через /reg! ✅")
            except Exception as e:
                logger.warning(f"Ошибка отправки уведомления пользователю {telegram_id}: {e}")
        logger.info("Рассылка о новом мероприятии завершена")
    except Exception as e:
        logger.error(f"Ошибка при рассылке уведомлений: {e}")

@admin_router.message(Command(commands=['stats_event']))
async def stats_event_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался получить статистику мероприятий")
        return
    try:
        events = get_all_events()
        for event in events:
            reg_count = get_registrations_count(event[0])
            donors_count = get_attended_count(event[0])
            keyboard = InlineKeyboardBuilder()
            keyboard.button(text="Заморозить ❄️" if event[5] == 'active' else "Разморозить 🔥",
                            callback_data=f"toggle_{event[0]}")
            await message.answer(f"Мероприятие: {event[1]} {event[2]} - {event[3]} 📅\n"
                                 f"Вместимость: {event[4]} 👥\nЗарегистрировано: {reg_count} 📝\nДоноров: {donors_count} 💉\nСтатус: {event[5]} ⚙️",
                                 reply_markup=keyboard.as_markup())
            logger.info(f"Админ {message.from_user.id} запросил статистику мероприятия ID {event[0]}")
    except Exception as e:
        logger.error(f"Ошибка при получении статистики мероприятий: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@admin_router.callback_query(lambda c: c.data.startswith('toggle_'))
async def toggle_event(callback_query: types.CallbackQuery):
    event_id = int(callback_query.data.split('_')[1])
    try:
        current_status = get_event_status(event_id)
        new_status = 'frozen' if current_status == 'active' else 'active'
        update_event_status(event_id, new_status)
        logger.info(f"Админ {callback_query.from_user.id} изменил статус мероприятия {event_id} на {new_status}")
        await callback_query.answer(f"Мероприятие {'заморожено ❄️' if new_status == 'frozen' else 'разморожено 🔥'}.")
    except Exception as e:
        logger.error(f"Ошибка при изменении статуса мероприятия {event_id}: {e}")
        await callback_query.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@admin_router.message(Command(commands=['see_profile']))
async def see_profile_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался просмотреть профили")
        return
    args = message.text.split()
    if len(args) > 1:
        try:
            search_id = int(args[1])
            await show_user_detail_by_id(message, search_id)
            return
        except ValueError:
            await message.answer("Некорректный ID. Используйте формат /see_profile <число>. ⚠️")
            logger.warning(f"Некорректный ID в поиске от админа {message.from_user.id}: {args[1]}")
            return
    await show_profiles(message, offset=0)

async def show_user_detail_by_id(message: types.Message, user_id: int):
    try:
        user = get_user_by_id(user_id)
        if not user:
            await message.answer("Пользователь не найден. ⚠️")
            logger.warning(f"Пользователь ID {user_id} не найден для админа {message.from_user.id}")
            return
        response = (f"Полная анкета: 📋\n"
                    f"ID: {user[0]}\n"
                    f"Telegram ID: {user[1]}\n"
                    f"Телефон: {user[2]} 📞\n"
                    f"ФИО: {user[3]}\n"
                    f"Категория: {user[4]}\n"
                    f"Группа: {user[5]} 📚\n"
                    f"Соцсети: {user[6]} 🔗\n"
                    f"DKM: {'Да' if user[7] else 'Нет'} 🦴\n"
                    f"Статус: {user[9]} ⚙️")
        keyboard = InlineKeyboardBuilder()
        keyboard.button(text="Кикнуть ❌", callback_data=f"kick_{user_id}")
        await message.answer(response, reply_markup=keyboard.as_markup())
        logger.info(f"Админ {message.from_user.id} запросил детали профиля по ID {user_id}")
    except Exception as e:
        logger.error(f"Ошибка при получении деталей профиля ID {user_id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

async def show_profiles(message: types.Message, offset: int):
    try:
        users = get_users_paginated(5, offset)
        if not users:
            await message.answer("Профили не найдены. 📭")
            logger.info("Профили пользователей не найдены")
            return
        pagination_keyboard = InlineKeyboardBuilder()
        if offset > 0:
            pagination_keyboard.button(text="Назад ⬅️", callback_data=f"prev_{offset - 5}")
        if len(users) == 5:
            pagination_keyboard.button(text="Вперед ➡️", callback_data=f"next_{offset + 5}")
        pagination_markup = pagination_keyboard.as_markup() if pagination_keyboard.inline_keyboard else None
        for user in users:
            reg_count = get_user_registrations_count(user[0])
            text = f"{user[1]}, Группа: {user[2]}, (ID: {user[0]}), Регистраций: {reg_count} 📝"
            keyboard = InlineKeyboardBuilder()
            keyboard.button(text="Подробнее 🔍", callback_data=f"detail_{user[0]}")
            await message.answer(text, reply_markup=keyboard.as_markup())
        if pagination_markup:
            await message.answer("Навигация: ↔️", reply_markup=pagination_markup)
        logger.info(f"Админ {message.from_user.id} запросил список профилей, offset: {offset}")
    except Exception as e:
        logger.error(f"Ошибка при получении списка профилей: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@admin_router.callback_query(lambda c: c.data.startswith('detail_'))
async def show_user_detail(callback_query: types.CallbackQuery):
    user_id = int(callback_query.data.split('_')[1])
    try:
        user = get_user_by_id(user_id)
        if not user:
            await callback_query.answer("Пользователь не найден. ⚠️")
            logger.warning(f"Пользователь ID {user_id} не найден")
            return
        response = (f"Полная анкета: 📋\n"
                    f"ID: {user[0]}\n"
                    f"Telegram ID: {user[1]}\n"
                    f"Телефон: {user[2]} 📞\n"
                    f"ФИО: {user[3]}\n"
                    f"Категория: {user[4]}\n"
                    f"Группа: {user[5]} 📚\n"
                    f"Соцсети: {user[6]} 🔗\n"
                    f"DKM: {'Да' if user[7] else 'Нет'} 🦴\n"
                    f"Статус: {user[9]} ⚙️")
        keyboard = InlineKeyboardBuilder()
        keyboard.button(text="Кикнуть ❌", callback_data=f"kick_{user_id}")
        await callback_query.message.answer(response, reply_markup=keyboard.as_markup())
        logger.info(f"Админ {callback_query.from_user.id} запросил детали профиля пользователя ID {user_id}")
    except Exception as e:
        logger.error(f"Ошибка при получении деталей профиля ID {user_id}: {e}")
        await callback_query.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@admin_router.callback_query(lambda c: c.data.startswith('kick_'))
async def kick_user(callback_query: types.CallbackQuery):
    # Локальный импорт bot внутри функции для избежания циклического импорта
    from src.bot import bot
    user_id = int(callback_query.data.split('_')[1])
    try:
        telegram_id = get_telegram_id_by_user_id(user_id)
        delete_user_by_id(user_id)
        if telegram_id is not None:
            await bot.send_message(telegram_id, "Ваш профиль был удален администратором. ❌")
        logger.info(f"Админ {callback_query.from_user.id} удалил пользователя ID {user_id}")
        await callback_query.answer("Пользователь удален. ✅")
    except Exception as e:
        logger.error(f"Ошибка при удалении пользователя ID {user_id}: {e}")
        await callback_query.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@admin_router.callback_query(lambda c: c.data.startswith('next_') or c.data.startswith('prev_'))
async def process_pagination(callback_query: types.CallbackQuery):
    action, offset = callback_query.data.split('_')
    offset = int(offset)
    await show_profiles(callback_query.message, offset)
    await callback_query.answer()
    logger.info(f"Админ {callback_query.from_user.id} запросил пагинацию профилей, offset: {offset}")

@admin_router.message(Command(commands=['import_excel']))
async def import_excel_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался выполнить импорт Excel")
        return
    try:
        import_from_excel()
        await message.answer("Данные из Excel успешно импортированы. ✅")
        logger.info(f"Админ {message.from_user.id} выполнил импорт данных из Excel")
    except Exception as e:
        await message.answer(f"Ошибка при импорте данных: {str(e)} ⚠️")
        logger.error(f"Ошибка при импорте Excel админом {message.from_user.id}: {e}")

@admin_router.message(Command(commands=['upload_stats']))
async def upload_stats_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    await message.answer("Отправьте файл Excel со статистикой (ФИО, дата, ЦК). 📂")

@admin_router.message(lambda message: message.document and message.document.file_name.endswith('.xlsx'))
async def process_upload_stats(message: types.Message):
    # Локальный импорт bot внутри функции для избежания циклического импорта
    from src.bot import bot
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    try:
        file_id = message.document.file_id
        file = await bot.get_file(file_id)
        file_path = file.file_path
        await bot.download_file(file_path, "temp_stats.xlsx")
        wb = openpyxl.load_workbook('temp_stats.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            fio = str(row[0]).strip()
            date = str(row[5]) if row[5] else None
            center = 'Гаврилова' if row[2] else 'ФМБА' if row[3] else None
            if not center or not date:
                continue
            user = get_user_by_fio(fio)
            if user:
                user_id = user[0]
                add_donation(user_id, date, center)
                if len(row) > 9 and row[9]:
                    update_dkm(user_id, 1)
        await message.answer("Статистика загружена и БД обновлена. ✅")
        logger.info(f"Админ {message.from_user.id} загрузил статистику из Excel")
    except Exception as e:
        await message.answer(f"Ошибка при загрузке: {str(e)} ⚠️")
        logger.error(f"Ошибка загрузки stats Excel: {e}")

@admin_router.message(Command(commands=['export_stats']))
async def export_stats_handler(message: types.Message):
    # Локальный импорт bot внутри функции для избежания циклического импорта
    from src.bot import bot
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['ID', 'ФИО', 'Группа', 'Кол-во Гаврилова', 'Кол-во ФМБА', 'Сумма', 'Последняя Гаврилова', 'Последняя ФМБА', 'Телефон'])
        users = get_all_users_for_export()
        for user in users:
            user_id = user[0]
            fio = user[3]
            group = user[5]
            phone = user[2]
            count_g = get_donations_count_by_center(user_id, "Гаврилова")
            count_f = get_donations_count_by_center(user_id, "ФМБА")
            sum_d = count_g + count_f
            last_g = get_last_donation(user_id)[0] if get_last_donation(user_id) and get_last_donation(user_id)[1] == "Гаврилова" else ''
            last_f = get_last_donation(user_id)[0] if get_last_donation(user_id) and get_last_donation(user_id)[1] == "ФМБА" else ''
            sheet.append([user_id, fio, group, count_g, count_f, sum_d, last_g, last_f, phone])
        wb.save('export_stats.xlsx')
        await bot.send_document(message.chat.id, types.FSInputFile('export_stats.xlsx'))
        logger.info(f"Админ {message.from_user.id} выгрузил статистику")
    except Exception as e:
        await message.answer(f"Ошибка при выгрузке: {str(e)} ⚠️")
        logger.error(f"Ошибка выгрузки stats: {e}")

# Функция регистрации хендлеров
def register_admin_handlers(dp: Dispatcher):
    dp.include_router(admin_router)