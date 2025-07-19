# bot.py
import asyncio
import schedule
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardMarkup, KeyboardButton
import re
import openpyxl  # Для export_stats, но загрузка в db.py

from db import (init_db, import_from_excel, get_user_by_phone, get_consent_by_phone, update_consent_by_phone,
                save_or_update_user, get_profile_status_by_telegram_id, get_active_events, get_user_by_telegram_id,
                get_donations_count_by_center, get_last_donation, get_donations_history, get_user_registrations,
                get_user_registrations_count, get_admin_stats, get_pending_users,
                update_profile_status, get_telegram_id_by_user_id, add_event, get_consented_users_telegram_ids,
                get_all_events, get_registrations_count, get_attended_count, get_event_status, update_event_status,
                get_user_by_id, get_users_paginated, delete_user_by_id, get_all_users_for_export,
                get_reminders_to_send, get_event_by_id, delete_reminder, get_past_events,
                get_non_attended_registrations,
                add_non_attendance_reason, get_user_id_by_telegram_id, get_event_capacity,
                get_registrations_count as get_event_reg_count,
                get_event_date, add_registration, add_reminder, cancel_registration, add_donation, update_dkm,
                get_user_by_name_surname, logger, add_question, get_unanswered_questions, mark_question_answered,
                get_user_telegram_id, get_connection)

# Токен бота
TOKEN = "7893139526:AAEw3mRwp8btOI4HWWhbLzL0j48kaQBUa50"
bot = Bot(token=TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# Функция проверки администратора (добавлен новый ID 1191457973)
def is_admin(user_id):
    admins = [123456789, 1653833795, 1191457973]  # Добавлен новый администратор
    return user_id in admins

# Классы состояний (адаптировано: убраны blood_group, medical_exemption, student_id)
class ProfilRegStates(StatesGroup):
    phone_confirm = State()
    name = State()
    surname = State()
    category = State()
    group = State()
    social_contacts = State()  # Добавлено для соцсетей (optional)

class ConsentStates(StatesGroup):
    consent = State()

class ProfilEditStates(StatesGroup):
    field = State()
    value = State()

class AddEventStates(StatesGroup):
    date = State()
    time = State()
    location = State()
    description = State()
    capacity = State()

class AskQuestionState(StatesGroup):
    text = State()

class AnswerQuestionState(StatesGroup):
    select = State()
    response = State()

class CancelReasonState(StatesGroup):
    reason = State()

class BroadcastState(StatesGroup):
    text = State()
    photo = State()
    confirm = State()

# --- Команды пользователя ---

@dp.message(Command(commands=['start']))
async def start_handler(message: types.Message, state: FSMContext):
    logger.info(f"Пользователь {message.from_user.id} вызвал команду /start")
    # Отправка баннера (изображения frame 3.jpg)
    try:
        await bot.send_photo(message.chat.id, types.FSInputFile('frame 3.jpg'), caption="Добро пожаловать в бот Дня Донора МИФИ! 💉❤️")
    except FileNotFoundError:
        logger.warning("Файл 'frame 3.jpg' не найден, баннер не отправлен")
        await message.answer("Добро пожаловать в бот Дня Донора МИФИ! 💉❤️")
    # Отправка PDF с политикой конфиденциальности
    try:
        await bot.send_document(message.chat.id, types.FSInputFile('privacy_policy.pdf'), caption="Ознакомьтесь с пользовательским соглашением (политика конфиденциальности). 📄")
    except FileNotFoundError:
        logger.warning("Файл 'privacy_policy.pdf' не найден")
        await message.answer("Файл с пользовательским соглашением не найден. Обратитесь к администратору. ⚠️")
        return
    except Exception as e:
        logger.error(f"Ошибка при отправке PDF: {e}")
        await message.answer("Произошла ошибка при отправке соглашения. Попробуйте позже. ⚠️")
        return
    # Запрос согласия с новой формулировкой
    await state.set_state(ConsentStates.consent)
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="Согласен ✅", callback_data="consent_yes")
    keyboard.button(text="Нет ❌", callback_data="consent_no")
    await message.answer("Я ознакомился с пользовательским соглашением и обязуюсь прочитать информацию из /info перед регистрацией на любое мероприятие. 📄",
                         reply_markup=keyboard.as_markup())

@dp.callback_query(lambda c: c.data in ['consent_yes', 'consent_no'], ConsentStates.consent)
async def process_initial_consent(callback_query: types.CallbackQuery, state: FSMContext):
    if callback_query.data == 'consent_yes':
        await state.update_data(initial_consent=True)  # Флаг согласия для новых пользователей
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Поделиться номером телефона 📞", request_contact=True)]],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await callback_query.message.answer("Спасибо! Теперь поделитесь номером телефона для авторизации. 🔑", reply_markup=keyboard)
        await state.set_state(ProfilRegStates.phone_confirm)
        logger.info(f"Пользователь {callback_query.from_user.id} принял согласие (ознакомление с PDF и /info)")
    else:
        await callback_query.message.answer("Без согласия бот не может работать. До свидания. 👋")
        await state.clear()
        logger.info(f"Пользователь {callback_query.from_user.id} отказался от согласия")
    await callback_query.answer()

@dp.message(ProfilRegStates.phone_confirm)
async def process_phone(message: types.Message, state: FSMContext):
    if not message.contact:
        await message.answer("Пожалуйста, используйте кнопку для отправки контакта. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} не отправил контакт")
        return
    phone = message.contact.phone_number
    try:
        user = get_user_by_phone(phone)
        await state.update_data(phone=phone)
        if user:
            await state.update_data(name=user[3], surname=user[4], category=user[5], group=user[6], social_contacts=user[7])  # Исправлен порядок surname/name
            response = f"Вы уже в базе: {user[3]} {user[4]}, категория: {user[5]}, группа: {user[6]}. Подтверждаете? (Да/Нет) ✅/❌"
            keyboard = InlineKeyboardBuilder()
            keyboard.button(text="Да ✅", callback_data="confirm_yes")
            keyboard.button(text="Нет ❌", callback_data="confirm_no")
            await message.answer(response, reply_markup=keyboard.as_markup())
            logger.info(f"Найден существующий пользователь с телефоном {phone}")
        else:
            await profil_reg_handler(message, state)
    except Exception as e:
        logger.error(f"Ошибка при проверке телефона {phone}: {e}")
        await message.answer("Произошла ошибка при проверке телефона. Попробуйте позже. ⚠️")

@dp.callback_query(lambda c: c.data in ['confirm_yes', 'confirm_no'])
async def confirm_existing(callback_query: types.CallbackQuery, state: FSMContext):
    if callback_query.data == 'confirm_yes':
        data = await state.get_data()
        phone = data['phone']
        try:
            consent = get_consent_by_phone(phone)
            if consent is None or not consent:
                # Для импортированных (consent=0) или новых — отправляем PDF и запрашиваем согласие
                try:
                    await bot.send_document(callback_query.message.chat.id, types.FSInputFile('privacy_policy.pdf'), caption="Ознакомьтесь с пользовательским соглашением (политика конфиденциальности). 📄")
                except FileNotFoundError:
                    logger.warning("Файл 'privacy_policy.pdf' не найден")
                    await callback_query.message.answer("Файл с пользовательским соглашением не найден. Обратитесь к администратору. ⚠️")
                    return
                except Exception as e:
                    logger.error(f"Ошибка при отправке PDF: {e}")
                    await callback_query.message.answer("Произошла ошибка при отправке соглашения. Попробуйте позже. ⚠️")
                    return
                await state.set_state(ConsentStates.consent)
                keyboard = InlineKeyboardBuilder()
                keyboard.button(text="Согласен ✅", callback_data="consent_yes")
                keyboard.button(text="Нет ❌", callback_data="consent_no")
                await callback_query.message.answer("Я ознакомился с пользовательским соглашением и обязуюсь прочитать информацию из /info перед регистрацией на любое мероприятие. 📄",
                                                    reply_markup=keyboard.as_markup())
                logger.info(f"Пользователь {phone} должен подтвердить согласие (с PDF и /info)")
            else:
                await callback_query.message.answer("Авторизация успешна! 🎉")
                # Отправка /help после успешной авторизации
                help_text = ("Вот команды пользователя: 📋\n"
                             "/profilReg - Зарегистрировать профиль ✍️\n"
                             "/reg - Записаться на мероприятие 📅\n"
                             "/profil - Посмотреть/изменить профиль 👤\n"
                             "/stats - Моя статистика 📊\n"
                             "/info - Информационные разделы 📖\n"
                             "/ask - Задать вопрос организаторам ❓\n"
                             "/help - Показать этот список ❓")
                await callback_query.message.answer(help_text)
                await state.clear()
                logger.info(f"Успешная авторизация пользователя {phone}")
        except Exception as e:
            logger.error(f"Ошибка при проверке согласия для телефона {phone}: {e}")
            await callback_query.message.answer("Произошла ошибка. Попробуйте позже. ⚠️")
    else:
        await profil_reg_handler(callback_query.message, state)
    await callback_query.answer()

@dp.callback_query(lambda c: c.data in ['consent_yes', 'consent_no'], ConsentStates.consent)
async def process_consent(callback_query: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    phone = data.get('phone')
    try:
        if callback_query.data == 'consent_yes':
            if phone:
                update_consent_by_phone(phone, 1)
            await callback_query.message.answer("Согласие принято! Авторизация успешна. 🎉")
            logger.info(f"Пользователь {phone or callback_query.from_user.id} принял согласие (ознакомление с PDF и /info)")
        else:
            await callback_query.message.answer("Без согласия бот не может работать. До свидания. 👋")
            logger.info(f"Пользователь {phone or callback_query.from_user.id} отказался от согласия")
        await state.clear()
    except Exception as e:
        logger.error(f"Ошибка при обновлении согласия для телефона {phone}: {e}")
        await callback_query.message.answer("Произошла ошибка. Попробуйте позже. ⚠️")
    await callback_query.answer()

@dp.message(Command(commands=['profilReg']))
async def profil_reg_handler(message: types.Message, state: FSMContext):
    await state.set_state(ProfilRegStates.name)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите ваше имя (только буквы): ✍️", reply_markup=keyboard)
    logger.info(f"Пользователь {message.from_user.id} начал регистрацию профиля")

@dp.message(ProfilRegStates.name)
async def process_name(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.clear()
        await message.answer("Регистрация отменена.", reply_markup=types.ReplyKeyboardRemove())
        return
    name = message.text.strip().capitalize()
    if not re.match(r'^[А-Яа-яA-Za-z\s]+$', name):
        await message.answer("Имя должно содержать только буквы. Попробуйте снова. ⚠️")
        logger.warning(f"Некорректное имя от пользователя {message.from_user.id}: {name}")
        return
    await state.update_data(name=name)
    await state.set_state(ProfilRegStates.surname)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите вашу фамилию (только буквы): ✍️", reply_markup=keyboard)

@dp.message(ProfilRegStates.surname)
async def process_surname(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(ProfilRegStates.name)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите ваше имя (только буквы): ✍️", reply_markup=keyboard)
        return
    surname = message.text.strip().capitalize()
    if not re.match(r'^[А-Яа-яA-Za-z\s]+$', surname):
        await message.answer("Фамилия должна содержать только буквы. Попробуйте снова. ⚠️")
        logger.warning(f"Некорректная фамилия от пользователя {message.from_user.id}: {surname}")
        return
    await state.update_data(surname=surname)
    await state.set_state(ProfilRegStates.category)
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="Студент 🎓", callback_data="cat_student")
    keyboard.button(text="Сотрудник 👔", callback_data="cat_employee")
    keyboard.button(text="Внешний донор 🌍", callback_data="cat_external")
    await message.answer("Выберите категорию: 📂", reply_markup=keyboard.as_markup())

@dp.callback_query(lambda c: c.data.startswith('cat_'))
async def process_category(callback_query: types.CallbackQuery, state: FSMContext):
    category = callback_query.data.split('_')[1]
    await state.update_data(category=category)
    if category == 'student':
        await state.set_state(ProfilRegStates.group)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await callback_query.message.answer("Введите номер группы (формат: Б21-302): 📚", reply_markup=keyboard)
    else:
        await state.set_state(ProfilRegStates.social_contacts)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await callback_query.message.answer("Введите контакты в соцсетях (или 'нет'): 🔗", reply_markup=keyboard)
    await callback_query.answer()
    logger.info(f"Пользователь {callback_query.from_user.id} выбрал категорию: {category}")

@dp.message(ProfilRegStates.group)
async def process_group(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(ProfilRegStates.category)
        keyboard = InlineKeyboardBuilder()
        keyboard.button(text="Студент 🎓", callback_data="cat_student")
        keyboard.button(text="Сотрудник 👔", callback_data="cat_employee")
        keyboard.button(text="Внешний донор 🌍", callback_data="cat_external")
        await message.answer("Выберите категорию: 📂", reply_markup=keyboard.as_markup())
        return
    group = message.text.strip().upper()
    if not re.match(r'^[А-Я]\d{2}-\d{3}$', group):
        await message.answer("Неверный формат группы (пример: Б21-302). Попробуйте снова. ⚠️")
        logger.warning(f"Некорректный формат группы от пользователя {message.from_user.id}: {group}")
        return
    await state.update_data(group=group)
    await state.set_state(ProfilRegStates.social_contacts)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите контакты в соцсетях (или 'нет'): 🔗", reply_markup=keyboard)

@dp.message(ProfilRegStates.social_contacts)
async def process_social_contacts(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        data = await state.get_data()
        category = data.get('category')
        if category == 'student':
            await state.set_state(ProfilRegStates.group)
            keyboard = ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="Назад 🔙")]],
                resize_keyboard=True
            )
            await message.answer("Введите номер группы (формат: Б21-302): 📚", reply_markup=keyboard)
        else:
            await state.set_state(ProfilRegStates.category)
            keyboard = InlineKeyboardBuilder()
            keyboard.button(text="Студент 🎓", callback_data="cat_student")
            keyboard.button(text="Сотрудник 👔", callback_data="cat_employee")
            keyboard.button(text="Внешний донор 🌍", callback_data="cat_external")
            await message.answer("Выберите категорию: 📂", reply_markup=keyboard.as_markup())
        return
    social_contacts = message.text.strip() if message.text.strip().lower() != 'нет' else None
    data = await state.get_data()
    try:
        save_or_update_user(message.from_user.id, data.get('phone'), data['name'], data['surname'],
                            data['category'], data.get('group'), social_contacts)
        await state.clear()
        await message.answer("Ваш профиль отправлен на модерацию. ⏳", reply_markup=types.ReplyKeyboardRemove())
    except Exception as e:
        logger.error(f"Ошибка при сохранении/обновлении профиля пользователя {data.get('name', 'Unknown')}: {e}")
        await message.answer("Произошла ошибка при сохранении профиля. Попробуйте позже. ⚠️")

@dp.message(Command(commands=['help']))
async def help_handler(message: types.Message):
    await message.answer("Вот команды пользователя:\n"
                         "/profilReg - Зарегистрировать профиль\n"
                         "/reg - Записаться на мероприятие\n"
                         "/profil - Посмотреть/изменить профиль\n"
                         "/stats - Моя статистика\n"
                         "/info - Информационные разделы\n"
                         "/ask - Задать вопрос организаторам ❓\n"
                         "/help - Показать этот список")
    logger.info(f"Пользователь {message.from_user.id} вызвал команду /help")

@dp.message(Command(commands=['reg']))
async def reg_handler(message: types.Message):
    try:
        profile_status = get_profile_status_by_telegram_id(message.from_user.id)
        if not profile_status or profile_status != 'approved':
            await message.answer("Ваш профиль не одобрен или не существует. ⚠️")
            logger.warning(f"Пользователь {message.from_user.id} не имеет одобренного профиля")
            return
        events = get_active_events()
        if not events:
            await message.answer("Нет доступных мероприятий. 📅")
            logger.info("Нет доступных мероприятий для регистрации")
            return
        keyboard = InlineKeyboardBuilder()
        for event in events:
            keyboard.button(text=f"{event[1]} {event[2]} - {event[4]} 📆", callback_data=f"reg_{event[0]}")
        await message.answer("Выберите мероприятие: 📋", reply_markup=keyboard.as_markup())
        logger.info(f"Пользователь {message.from_user.id} запросил регистрацию на мероприятие")
    except Exception as e:
        logger.error(f"Ошибка при получении мероприятий: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.callback_query(lambda c: c.data.startswith('reg_'))
async def process_register(callback_query: types.CallbackQuery):
    event_id = int(callback_query.data.split('_')[1])
    user_id = callback_query.from_user.id
    try:
        db_user_id = get_user_id_by_telegram_id(user_id)
        capacity = get_event_capacity(event_id)
        registered_count = get_event_reg_count(event_id)
        if registered_count >= capacity:
            await callback_query.answer("Мероприятие заполнено. ❌")
            logger.warning(f"Мероприятие {event_id} заполнено")
            return
        event_date = get_event_date(event_id)
        try:
            reminder_date = (datetime.strptime(event_date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
        except ValueError:
            await callback_query.answer("Ошибка: некорректная дата мероприятия. ⚠️")
            logger.error(f"Некорректная дата мероприятия {event_id}: {event_date}")
            return
        add_registration(db_user_id, event_id)
        add_reminder(db_user_id, event_id, reminder_date)
        logger.info(f"Пользователь {user_id} зарегистрирован на мероприятие {event_id}")
        await callback_query.answer("Вы зарегистрированы! ✅")
    except Exception as e:
        logger.error(f"Ошибка при регистрации пользователя {user_id} на мероприятие {event_id}: {e}")
        await callback_query.answer("Произошла ошибка при регистрации. Попробуйте позже. ⚠️")

@dp.message(Command(commands=['profil']))
async def profil_handler(message: types.Message, state: FSMContext):
    try:
        user = get_user_by_telegram_id(message.from_user.id)
        if not user:
            await message.answer("Профиль не найден. ⚠️")
            logger.warning(f"Профиль пользователя {message.from_user.id} не найден")
            return
        user_id = user[0]
        # Вычисление статистики из donations
        count_gavrilov = get_donations_count_by_center(user_id, "Гаврилова")
        count_fmba = get_donations_count_by_center(user_id, "ФМБА")
        sum_donations = count_gavrilov + count_fmba
        last_donation = get_last_donation(user_id)
        last_date_center = f"{last_donation[0]} / {last_donation[1]}" if last_donation else "Нет"
        history = get_donations_history(user_id)
        history_str = "\n".join([f"{d[0]} - {d[1]}" for d in history]) if history else "Нет истории"
        dkm_str = "Да" if user[6] else "Нет"
        response = (
            f"Ваш профиль: 📋\nИмя: {user[1]}\nФамилия: {user[2]}\nКатегория: {user[3]}\nГруппа: {user[4]}\n"
            f"Соцсети: {user[5] or 'Нет'} 🔗\nСтатус: {user[7]} ⚙️\n"
            f"Количество донаций: {sum_donations} 💉\nПоследняя донация: {last_date_center} 📅\n"
            f"Вступление в ДКМ: {dkm_str} 🦴\nИстория донаций:\n{history_str}")
        # Добавляем список текущих регистраций с кнопками отмены
        registrations = get_user_registrations(user_id)
        if registrations:
            response += "\n\nВаши текущие регистрации: 📅"
            keyboard = InlineKeyboardBuilder()
            for reg in registrations:
                response += f"\n- {reg[1]} {reg[2]} - {reg[3]}"
                keyboard.button(text=f"Отменить {reg[1]} ❌", callback_data=f"unreg_{reg[0]}")
            await message.answer(response, reply_markup=keyboard.as_markup())
        else:
            await message.answer(response)
        logger.info(f"Пользователь {message.from_user.id} запросил просмотр профиля")
    except Exception as e:
        logger.error(f"Ошибка при получении профиля пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.callback_query(lambda c: c.data.startswith('unreg_'))
async def process_unreg(callback_query: types.CallbackQuery, state: FSMContext):
    event_id = int(callback_query.data.split('_')[1])
    user_id = callback_query.from_user.id
    try:
        db_user_id = get_user_id_by_telegram_id(user_id)
        cancel_registration(db_user_id, event_id)
        await state.set_state(CancelReasonState.reason)
        await state.update_data(reg_id=get_registration_id(db_user_id, event_id))  # Получить reg_id
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Медотвод ⚕️")],
                      [KeyboardButton(text="Личные причины 👤")],
                      [KeyboardButton(text="Не захотел 😔")]],
            resize_keyboard=True
        )
        await callback_query.message.answer("Регистрация отменена. Пожалуйста, укажите причину отмены:", reply_markup=keyboard)
        logger.info(f"Пользователь {user_id} отменил регистрацию на мероприятие {event_id}, запрошенная причина")
    except Exception as e:
        logger.error(f"Ошибка при отмене регистрации пользователя {user_id} на {event_id}: {e}")
        await callback_query.answer("Произошла ошибка. Попробуйте позже. ⚠️")

def get_registration_id(user_id, event_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM registrations WHERE user_id = ? AND event_id = ?', (user_id, event_id))
        result = cursor.fetchone()
        return result[0] if result else None

@dp.message(CancelReasonState.reason)
async def process_cancel_reason(message: types.Message, state: FSMContext):
    reason = message.text
    data = await state.get_data()
    reg_id = data.get('reg_id')
    if reg_id:
        add_non_attendance_reason(reg_id, reason)
        await message.answer("Причина отмены записана. Спасибо!", reply_markup=types.ReplyKeyboardRemove())
        logger.info(f"Записана причина отмены для reg {reg_id}: {reason}")
    await state.clear()

@dp.message(Command(commands=['stats']))
async def stats_handler(message: types.Message):
    try:
        user_id = get_user_id_by_telegram_id(message.from_user.id)
        if user_id:
            reg_count = get_user_registrations_count(user_id)
            await message.answer(f"Ваша статистика: 📊\nЗарегистрировано на мероприятий: {reg_count} 📅")
            logger.info(f"Пользователь {message.from_user.id} запросил статистику: {reg_count} регистраций")
        else:
            await message.answer("Вы не зарегистрированы. ⚠️")
            logger.warning(f"Пользователь {message.from_user.id} не зарегистрирован")
    except Exception as e:
        logger.error(f"Ошибка при получении статистики пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.message(Command(commands=['info']))
async def info_handler(message: types.Message):
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="О донорстве крови", callback_data="info_blood")
    keyboard.button(text="О донорстве костного мозга", callback_data="info_bone")
    keyboard.button(text="О донациях в МИФИ", callback_data="info_mifi")
    await message.answer("Выберите раздел информации: 📖", reply_markup=keyboard.as_markup())
    logger.info(f"Пользователь {message.from_user.id} запросил информационные разделы")

@dp.callback_query(lambda c: c.data.startswith('info_'))
async def process_info(callback_query: types.CallbackQuery):
    section_map = {
        'blood': 'blood_donation.txt',
        'bone': 'bone_marrow_donation.txt',
        'mifi': 'mifi_donations.txt'
    }
    file_name = section_map.get(callback_query.data.split('_')[1])
    if not file_name:
        await callback_query.answer("Некорректный раздел. ⚠️")
        return
    try:
        with open(file_name, 'r', encoding='utf-8') as f:
            text = f.read()
        await callback_query.message.answer(text)
        logger.info(f"Пользователь {callback_query.from_user.id} просмотрел раздел из {file_name}")
    except FileNotFoundError:
        await callback_query.message.answer("Текст раздела не найден. ⚠️")
        logger.warning(f"Файл {file_name} не найден")
    except Exception as e:
        logger.error(f"Ошибка при чтении файла {file_name}: {e}")
        await callback_query.answer("Произошла ошибка. Попробуйте позже. ⚠️")
    await callback_query.answer()

@dp.message(Command(commands=['ask']))
async def ask_handler(message: types.Message, state: FSMContext):
    profile_status = get_profile_status_by_telegram_id(message.from_user.id)
    if not profile_status or profile_status != 'approved':
        await message.answer("Ваш профиль не одобрен или не существует. ⚠️")
        return
    await state.set_state(AskQuestionState.text)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите ваш вопрос или сообщение организаторам:", reply_markup=keyboard)

@dp.message(AskQuestionState.text)
async def process_ask_text(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.clear()
        await message.answer("Операция отменена.", reply_markup=types.ReplyKeyboardRemove())
        return
    text = message.text.strip()
    if not text:
        await message.answer("Сообщение не может быть пустым. Попробуйте снова.")
        return
    try:
        user_id = get_user_id_by_telegram_id(message.from_user.id)
        question_id = add_question(user_id, text)
        await message.answer("Ваш вопрос отправлен организаторам. Они ответят в ближайшее время.", reply_markup=types.ReplyKeyboardRemove())
        # Уведомление админам с обработкой ошибок
        admins = [123456789, 1653833795, 1191457973]
        for admin_id in admins:
            try:
                await bot.send_message(admin_id, f"Новый вопрос от пользователя ID {user_id}: {text}\nОтветьте через /answer")
            except Exception as e:
                logger.error(f"Ошибка при отправке уведомления админу {admin_id}: {e}")
        logger.info(f"Пользователь {message.from_user.id} задал вопрос: {text}")
    except Exception as e:
        logger.error(f"Ошибка при отправке вопроса от {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже.")
    await state.clear()

# --- Админские команды ---

@dp.message(Command(commands=['answer']))
async def answer_handler(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    try:
        questions = get_unanswered_questions()
        if not questions:
            await message.answer("Нет неотвеченных вопросов.")
            return
        keyboard = InlineKeyboardBuilder()
        for q in questions:
            user_tg_id = get_user_telegram_id(q[1])
            keyboard.button(text=f"Вопрос {q[0]} от {user_tg_id}", callback_data=f"ans_{q[0]}")
        await message.answer("Выберите вопрос для ответа:", reply_markup=keyboard.as_markup())
        await state.set_state(AnswerQuestionState.select)
    except Exception as e:
        logger.error(f"Ошибка при получении вопросов для ответа: {e}")
        await message.answer("Произошла ошибка.")

@dp.callback_query(lambda c: c.data.startswith('ans_'), AnswerQuestionState.select)
async def select_question(callback_query: types.CallbackQuery, state: FSMContext):
    question_id = int(callback_query.data.split('_')[1])
    await state.update_data(question_id=question_id)
    await state.set_state(AnswerQuestionState.response)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await callback_query.message.answer("Введите ответ на вопрос:", reply_markup=keyboard)
    await callback_query.answer()

@dp.message(AnswerQuestionState.response)
async def process_answer_text(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(AnswerQuestionState.select)
        await answer_handler(message, state)  # Вернуться к списку
        return
    text = message.text.strip()
    if not text:
        await message.answer("Ответ не может быть пустым. Попробуйте снова.")
        return
    data = await state.get_data()
    question_id = data.get('question_id')
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT user_id FROM questions WHERE id = ?', (question_id,))
            user_id = cursor.fetchone()[0]
        user_tg_id = get_user_telegram_id(user_id)
        if user_tg_id:
            await bot.send_message(user_tg_id, f"Ответ на ваш вопрос от организаторов: {text}")
            mark_question_answered(question_id)
            await message.answer("Ответ отправлен пользователю.", reply_markup=types.ReplyKeyboardRemove())
            logger.info(f"Админ {message.from_user.id} ответил на вопрос {question_id}")
        else:
            await message.answer("Не удалось найти Telegram ID пользователя.")
    except Exception as e:
        logger.error(f"Ошибка при отправке ответа на вопрос {question_id}: {e}")
        await message.answer("Произошла ошибка.")
    await state.clear()

@dp.message(Command(commands=['broadcast']))
async def broadcast_handler(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    await state.set_state(BroadcastState.text)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите текст для рассылки всем пользователям:", reply_markup=keyboard)

@dp.message(BroadcastState.text)
async def process_broadcast_text(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.clear()
        await message.answer("Рассылка отменена.", reply_markup=types.ReplyKeyboardRemove())
        return
    text = message.text.strip()
    if not text:
        await message.answer("Текст не может быть пустым. Попробуйте снова.")
        return
    await state.update_data(text=text)
    await state.set_state(BroadcastState.photo)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Без фото 📄")],
                  [KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Прикрепите фото (если нужно) или нажмите 'Без фото':", reply_markup=keyboard)

@dp.message(BroadcastState.photo)
async def process_broadcast_photo(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(BroadcastState.text)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите текст для рассылки:", reply_markup=keyboard)
        return
    if message.text == "Без фото 📄":
        photo = None
    elif message.photo:
        photo = message.photo[-1].file_id
    else:
        await message.answer("Пожалуйста, прикрепите фото или выберите 'Без фото'.")
        return
    await state.update_data(photo=photo)
    await state.set_state(BroadcastState.confirm)
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="Подтвердить ✅", callback_data="broadcast_confirm")
    keyboard.button(text="Отмена ❌", callback_data="broadcast_cancel")
    await message.answer("Подтвердите рассылку?", reply_markup=keyboard.as_markup())

@dp.callback_query(lambda c: c.data in ['broadcast_confirm', 'broadcast_cancel'], BroadcastState.confirm)
async def confirm_broadcast(callback_query: types.CallbackQuery, state: FSMContext):
    if callback_query.data == 'broadcast_cancel':
        await state.clear()
        await callback_query.message.answer("Рассылка отменена.")
        await callback_query.answer()
        return
    data = await state.get_data()
    text = data.get('text')
    photo = data.get('photo')
    try:
        users = get_consented_users_telegram_ids()
        sent_count = 0
        for tg_id in users:
            try:
                if photo:
                    await bot.send_photo(tg_id, photo, caption=text)
                else:
                    await bot.send_message(tg_id, text)
                sent_count += 1
            except Exception as e:
                logger.warning(f"Ошибка отправки рассылки пользователю {tg_id}: {e}")
        await callback_query.message.answer(f"Рассылка завершена. Отправлено {sent_count} пользователям.")
        logger.info(f"Админ {callback_query.from_user.id} отправил рассылку: {text[:50]}...")
    except Exception as e:
        logger.error(f"Ошибка при рассылке: {e}")
        await callback_query.message.answer("Произошла ошибка при рассылке.")
    await state.clear()
    await callback_query.answer()

@dp.message(Command(commands=['admin_stats']))
async def admin_stats_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался получить админскую статистику")
        return
    try:
        users_count, events_count, reg_count = get_admin_stats()
        await message.answer(
            f"Статистика: 📊\nПользователей: {users_count} 👥\nМероприятий: {events_count} 📅\nРегистраций: {reg_count} 📝")
        logger.info(f"Админ {message.from_user.id} запросил статистику")
    except Exception as e:
        logger.error(f"Ошибка при получении админской статистики: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.message(Command(commands=['admin_reg']))
async def admin_reg_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался управлять заявками")
        return
    try:
        pending_users = get_pending_users()
        if not pending_users:
            await message.answer("Нет заявок. 📭")
            logger.info("Нет заявок на модерацию")
            return
        for user in pending_users:
            keyboard = InlineKeyboardBuilder()
            keyboard.button(text="Принять ✅", callback_data=f"approve_{user[0]}")
            keyboard.button(text="Отклонить ❌", callback_data=f"reject_{user[0]}")
            await message.answer(
                f"Заявка: {user[1]} {user[2]}\nГруппа: {user[3]}\nСоцсети: {user[4]}",
                reply_markup=keyboard.as_markup())
            logger.info(f"Отображена заявка пользователя {user[1]} {user[2]} для админа {message.from_user.id}")
    except Exception as e:
        logger.error(f"Ошибка при получении заявок на модерацию: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.callback_query(lambda c: c.data.startswith('approve_') or c.data.startswith('reject_'))
async def process_profile_action(callback_query: types.CallbackQuery):
    action, user_id_str = callback_query.data.split('_')
    user_id = int(user_id_str)
    try:
        status = 'approved' if action == 'approve' else 'rejected'
        update_profile_status(user_id, status)
        telegram_id = get_telegram_id_by_user_id(user_id)
        if action == 'approve' and telegram_id is not None:  # Фикс бага: проверка на None перед отправкой
            await bot.send_message(telegram_id, "Ваш профиль был принят администратором. ✅")
            # Отправка /help после approve
            help_text = ("Вот команды пользователя: 📋\n"
                         "/profilReg - Зарегистрировать профиль ✍️\n"
                         "/reg - Записаться на мероприятие 📅\n"
                         "/profil - Посмотреть/изменить профиль 👤\n"
                         "/stats - Моя статистика 📊\n"
                         "/info - Информационные разделы 📖\n"
                         "/ask - Задать вопрос организаторам ❓\n"
                         "/help - Показать этот список ❓")
            await bot.send_message(telegram_id, help_text)
        logger.info(f"Профиль пользователя ID {user_id} {status} админом {callback_query.from_user.id}")
        await callback_query.answer(f"Профиль {'принят ✅' if action == 'approve' else 'отклонен ❌'}.")
    except Exception as e:
        logger.error(f"Ошибка при обработке профиля ID {user_id}: {e}")
        await callback_query.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.message(Command(commands=['admin_help']))
async def admin_help_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался получить админскую помощь")
        return
    await message.answer("/admin_stats - Статистика проекта\n"
                         "/admin_reg - Управление заявками\n"
                         "/add_event - Добавить мероприятие\n"
                         "/stats_event - Статистика мероприятий\n"
                         "/see_profile - Просмотр профилей\n"
                         "/see_profile (числовый аргумент) - Просмотр конкретного пользователя по айди\n"
                         "/import_excel - Импорт из Excel\n"
                         "/upload_stats - Загрузить статистику из Excel\n"
                         "/export_stats - Выгрузить статистику в Excel\n"
                         "/answer - Ответить на вопросы пользователей\n"
                         "/broadcast - Рассылка сообщений всем пользователям\n"
                         "/help - Список пользовательских команд")
    logger.info(f"Админ {message.from_user.id} запросил список админских команд")

@dp.message(Command(commands=['add_event']))
async def add_event_handler(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался добавить мероприятие")
        return
    await state.set_state(AddEventStates.date)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите дату (YYYY-MM-DD): 📅", reply_markup=keyboard)
    logger.info(f"Админ {message.from_user.id} начал добавление мероприятия")

@dp.message(AddEventStates.date)
async def process_event_date(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.clear()
        await message.answer("Добавление отменено.", reply_markup=types.ReplyKeyboardRemove())
        return
    try:
        event_date = datetime.strptime(message.text, '%Y-%m-%d')
        if event_date < datetime.now().replace(hour=0, minute=0, second=0, microsecond=0):
            await message.answer("Дата уже прошла. Введите будущую дату в формате YYYY-MM-DD. ⚠️")
            logger.warning(f"Попытка создать мероприятие с прошедшей датой {message.text} от админа {message.from_user.id}")
            return
    except ValueError:
        await message.answer("Некорректный формат даты. Пожалуйста, введите дату в формате YYYY-MM-DD. ⚠️")
        logger.warning(f"Некорректный формат даты от админа {message.from_user.id}: {message.text}")
        return
    await state.update_data(date=message.text)
    await state.set_state(AddEventStates.time)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите время (HH:MM): ⏰", reply_markup=keyboard)

@dp.message(AddEventStates.time)
async def process_event_time(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(AddEventStates.date)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите дату (YYYY-MM-DD): 📅", reply_markup=keyboard)
        return
    await state.update_data(time=message.text)
    await state.set_state(AddEventStates.location)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите место: 📍", reply_markup=keyboard)

@dp.message(AddEventStates.location)
async def process_event_location(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(AddEventStates.time)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите время (HH:MM): ⏰", reply_markup=keyboard)
        return
    await state.update_data(location=message.text)
    await state.set_state(AddEventStates.description)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите описание: 📝", reply_markup=keyboard)

@dp.message(AddEventStates.description)
async def process_event_description(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(AddEventStates.location)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите место: 📍", reply_markup=keyboard)
        return
    await state.update_data(description=message.text)
    await state.set_state(AddEventStates.capacity)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Назад 🔙")]],
        resize_keyboard=True
    )
    await message.answer("Введите вместимость: 👥", reply_markup=keyboard)

@dp.message(AddEventStates.capacity)
async def process_event_capacity(message: types.Message, state: FSMContext):
    if message.text == "Назад 🔙":
        await state.set_state(AddEventStates.description)
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Назад 🔙")]],
            resize_keyboard=True
        )
        await message.answer("Введите описание: 📝", reply_markup=keyboard)
        return
    try:
        capacity = int(message.text)
        if capacity <= 0:
            raise ValueError("Вместимость должна быть положительным числом")
    except ValueError:
        await message.answer("Вместимость должна быть числом больше 0. ⚠️")
        logger.warning(f"Некорректная вместимость от админа {message.from_user.id}: {message.text}")
        return
    data = await state.get_data()
    try:
        add_event(data['date'], data['time'], data['location'], data['description'], capacity)
        await state.clear()
        await message.answer("Мероприятие добавлено. ✅", reply_markup=types.ReplyKeyboardRemove())
        # Рассылка уведомлений всем consented пользователям
        asyncio.create_task(send_new_event_notification(data['date'], data['time'], data['location'], data['description']))
        logger.info(f"Админ {message.from_user.id} добавил мероприятие: {data['description']}")
    except Exception as e:
        logger.error(f"Ошибка при добавлении мероприятия: {e}")
        await message.answer("Произошла ошибка при добавлении мероприятия. Попробуйте позже. ⚠️")

async def send_new_event_notification(date, time, location, description):
    try:
        users = get_consented_users_telegram_ids()
        for telegram_id in users:
            try:
                await bot.send_message(telegram_id, f"Новое мероприятие: {description} 📅\nДата: {date} {time} ⏰\nМесто: {location} 📍\nЗарегистрируйтесь через /reg! ✅")
            except Exception as e:
                logger.warning(f"Ошибка отправки уведомления пользователю {telegram_id}: {e}")
        logger.info("Рассылка о новом мероприятии завершена")
    except Exception as e:
        logger.error(f"Ошибка при рассылке уведомлений: {e}")

@dp.message(Command(commands=['stats_event']))
async def stats_event_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался получить статистику мероприятий")
        return
    try:
        events = get_all_events()
        for event in events:
            reg_count = get_registrations_count(event[0])
            donors_count = get_attended_count(event[0])
            keyboard = InlineKeyboardBuilder()
            keyboard.button(text="Заморозить ❄️" if event[5] == 'active' else "Разморозить 🔥",
                            callback_data=f"toggle_{event[0]}")
            await message.answer(f"Мероприятие: {event[1]} {event[2]} - {event[3]} 📅\n"
                                 f"Вместимость: {event[4]} 👥\nЗарегистрировано: {reg_count} 📝\nДоноров: {donors_count} 💉\nСтатус: {event[5]} ⚙️",
                                 reply_markup=keyboard.as_markup())
            logger.info(f"Админ {message.from_user.id} запросил статистику мероприятия ID {event[0]}")
    except Exception as e:
        logger.error(f"Ошибка при получении статистики мероприятий: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.callback_query(lambda c: c.data.startswith('toggle_'))
async def toggle_event(callback_query: types.CallbackQuery):
    event_id = int(callback_query.data.split('_')[1])
    try:
        current_status = get_event_status(event_id)
        new_status = 'frozen' if current_status == 'active' else 'active'
        update_event_status(event_id, new_status)
        logger.info(f"Админ {callback_query.from_user.id} изменил статус мероприятия {event_id} на {new_status}")
        await callback_query.answer(f"Мероприятие {'заморожено ❄️' if new_status == 'frozen' else 'разморожено 🔥'}.")
    except Exception as e:
        logger.error(f"Ошибка при изменении статуса мероприятия {event_id}: {e}")
        await callback_query.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.message(Command(commands=['see_profile']))
async def see_profile_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался просмотреть профили")
        return
    # Проверка на аргумент (поиск по ID)
    args = message.text.split()
    if len(args) > 1:
        try:
            search_id = int(args[1])
            await show_user_detail_by_id(message, search_id)
            return
        except ValueError:
            await message.answer("Некорректный ID. Используйте формат /see_profile <число>. ⚠️")
            logger.warning(f"Некорректный ID в поиске от админа {message.from_user.id}: {args[1]}")
            return
    # Если нет аргумента, показываем список с пагинацией
    await show_profiles(message, offset=0)

async def show_user_detail_by_id(message: types.Message, user_id: int):
    try:
        user = get_user_by_id(user_id)
        if not user:
            await message.answer("Пользователь не найден. ⚠️")
            logger.warning(f"Пользователь ID {user_id} не найден для админа {message.from_user.id}")
            return
        response = (f"Полная анкета: 📋\n"
                    f"ID: {user[0]}\n"
                    f"Telegram ID: {user[1]}\n"
                    f"Телефон: {user[2]} 📞\n"
                    f"Имя: {user[3]}\n"
                    f"Фамилия: {user[4]}\n"
                    f"Категория: {user[5]}\n"
                    f"Группа: {user[6]} 📚\n"
                    f"Соцсети: {user[7]} 🔗\n"
                    f"DKM: {'Да' if user[8] else 'Нет'} 🦴\n"
                    f"Статус: {user[10]} ⚙️")
        keyboard = InlineKeyboardBuilder()
        keyboard.button(text="Кикнуть ❌", callback_data=f"kick_{user_id}")
        await message.answer(response, reply_markup=keyboard.as_markup())
        logger.info(f"Админ {message.from_user.id} запросил детали профиля по ID {user_id}")
    except Exception as e:
        logger.error(f"Ошибка при получении деталей профиля ID {user_id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

async def show_profiles(message: types.Message, offset: int):
    try:
        users = get_users_paginated(5, offset)
        if not users:
            await message.answer("Профили не найдены. 📭")
            logger.info("Профили пользователей не найдены")
            return
        # Пагинационные кнопки
        pagination_keyboard = InlineKeyboardBuilder()
        if offset > 0:
            pagination_keyboard.button(text="Назад ⬅️", callback_data=f"prev_{offset - 5}")
        if len(users) == 5:
            pagination_keyboard.button(text="Вперед ➡️", callback_data=f"next_{offset + 5}")
        pagination_markup = pagination_keyboard.as_markup() if pagination_keyboard.inline_keyboard else None
        # Для каждого пользователя: текст + кнопка "Подробнее"
        for user in users:
            reg_count = get_user_registrations_count(user[0])
            text = f"{user[2]} {user[1]}, Группа: {user[3]}, (ID: {user[0]}), Регистраций: {reg_count} 📝"
            keyboard = InlineKeyboardBuilder()
            keyboard.button(text="Подробнее 🔍", callback_data=f"detail_{user[0]}")
            await message.answer(text, reply_markup=keyboard.as_markup())
        # Отправляем пагинационные кнопки отдельно, если есть
        if pagination_markup:
            await message.answer("Навигация: ↔️", reply_markup=pagination_markup)
        logger.info(f"Админ {message.from_user.id} запросил список профилей, offset: {offset}")
    except Exception as e:
        logger.error(f"Ошибка при получении списка профилей: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.callback_query(lambda c: c.data.startswith('detail_'))
async def show_user_detail(callback_query: types.CallbackQuery):
    user_id = int(callback_query.data.split('_')[1])
    try:
        user = get_user_by_id(user_id)
        if not user:
            await callback_query.answer("Пользователь не найден. ⚠️")
            logger.warning(f"Пользователь ID {user_id} не найден")
            return
        response = (f"Полная анкета: 📋\n"
                    f"ID: {user[0]}\n"
                    f"Telegram ID: {user[1]}\n"
                    f"Телефон: {user[2]} 📞\n"
                    f"Имя: {user[3]}\n"
                    f"Фамилия: {user[4]}\n"
                    f"Категория: {user[5]}\n"
                    f"Группа: {user[6]} 📚\n"
                    f"Соцсети: {user[7]} 🔗\n"
                    f"DKM: {'Да' if user[8] else 'Нет'} 🦴\n"
                    f"Статус: {user[10]} ⚙️")
        keyboard = InlineKeyboardBuilder()
        keyboard.button(text="Кикнуть ❌", callback_data=f"kick_{user_id}")
        await callback_query.message.answer(response, reply_markup=keyboard.as_markup())
        logger.info(f"Админ {callback_query.from_user.id} запросил детали профиля пользователя ID {user_id}")
    except Exception as e:
        logger.error(f"Ошибка при получении деталей профиля ID {user_id}: {e}")
        await callback_query.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.callback_query(lambda c: c.data.startswith('kick_'))
async def kick_user(callback_query: types.CallbackQuery):
    user_id = int(callback_query.data.split('_')[1])
    try:
        telegram_id = get_telegram_id_by_user_id(user_id)
        delete_user_by_id(user_id)
        if telegram_id is not None:  # Фикс: проверка на None
            await bot.send_message(telegram_id, "Ваш профиль был удален администратором. ❌")
        logger.info(f"Админ {callback_query.from_user.id} удалил пользователя ID {user_id}")
        await callback_query.answer("Пользователь удален. ✅")
    except Exception as e:
        logger.error(f"Ошибка при удалении пользователя ID {user_id}: {e}")
        await callback_query.answer("Произошла ошибка. Попробуйте позже. ⚠️")

@dp.callback_query(lambda c: c.data.startswith('next_') or c.data.startswith('prev_'))
async def process_pagination(callback_query: types.CallbackQuery):
    action, offset = callback_query.data.split('_')
    offset = int(offset)
    await show_profiles(callback_query.message, offset)
    await callback_query.answer()
    logger.info(f"Админ {callback_query.from_user.id} запросил пагинацию профилей, offset: {offset}")

@dp.message(Command(commands=['import_excel']))
async def import_excel_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        logger.warning(f"Пользователь {message.from_user.id} пытался выполнить импорт Excel")
        return
    try:
        import_from_excel()
        await message.answer("Данные из Excel успешно импортированы. ✅")
        logger.info(f"Админ {message.from_user.id} выполнил импорт данных из Excel")
    except Exception as e:
        await message.answer(f"Ошибка при импорте данных: {str(e)} ⚠️")
        logger.error(f"Ошибка при импорте Excel админом {message.from_user.id}: {e}")

@dp.message(Command(commands=['upload_stats']))
async def upload_stats_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    await message.answer("Отправьте файл Excel со статистикой (ФИО, дата, ЦК). 📂")

@dp.message(lambda message: message.document and message.document.file_name.endswith('.xlsx'))
async def process_upload_stats(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    try:
        file_id = message.document.file_id
        file = await bot.get_file(file_id)
        file_path = file.file_path
        await bot.download_file(file_path, "temp_stats.xlsx")
        wb = openpyxl.load_workbook('temp_stats.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            fio = str(row[0]).strip()
            parts = fio.split(maxsplit=1)
            surname = parts[0]
            name = parts[1] if len(parts) > 1 else ''
            date = str(row[5]) if row[5] else None  # Пример: дата Гаврилова или ФМБА
            center = 'Гаврилова' if row[2] else 'ФМБА' if row[3] else None
            if not center or not date:
                continue
            user = get_user_by_name_surname(name, surname)
            if user:
                user_id = user[0]
                add_donation(user_id, date, center)
                # Если DKM (допустим, добавим колонку 9 как DKM)
                if len(row) > 9 and row[9]:
                    update_dkm(user_id, 1)
        await message.answer("Статистика загружена и БД обновлена. ✅")
        logger.info(f"Админ {message.from_user.id} загрузил статистику из Excel")
    except Exception as e:
        await message.answer(f"Ошибка при загрузке: {str(e)} ⚠️")
        logger.error(f"Ошибка загрузки stats Excel: {e}")

@dp.message(Command(commands=['export_stats']))
async def export_stats_handler(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет прав. ⚠️")
        return
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['ID', 'ФИО', 'Группа', 'Кол-во Гаврилова', 'Кол-во ФМБА', 'Сумма', 'Последняя Гаврилова', 'Последняя ФМБА', 'Телефон'])
        users = get_all_users_for_export()
        for user in users:
            user_id = user[0]
            fio = f"{user[4]} {user[3]}"
            group = user[6]
            phone = user[2]
            count_g = get_donations_count_by_center(user_id, "Гаврилова")
            count_f = get_donations_count_by_center(user_id, "ФМБА")
            sum_d = count_g + count_f
            last_g = get_last_donation(user_id)[0] if get_last_donation(user_id) and get_last_donation(user_id)[1] == "Гаврилова" else ''
            last_f = get_last_donation(user_id)[0] if get_last_donation(user_id) and get_last_donation(user_id)[1] == "ФМБА" else ''
            sheet.append([user_id, fio, group, count_g, count_f, sum_d, last_g, last_f, phone])
        wb.save('export_stats.xlsx')
        await bot.send_document(message.chat.id, types.FSInputFile('export_stats.xlsx'))
        logger.info(f"Админ {message.from_user.id} выгрузил статистику")
    except Exception as e:
        await message.answer(f"Ошибка при выгрузке: {str(e)} ⚠️")
        logger.error(f"Ошибка выгрузки stats: {e}")

# --- Напоминания и опросы ---

async def check_reminders():
    try:
        current_date = datetime.now().strftime('%Y-%m-%d')
        reminders = get_reminders_to_send(current_date)
        for reminder in reminders:
            user_id = reminder[1]
            event_id = reminder[2]
            event = get_event_by_id(event_id)
            if event:
                await bot.send_message(user_id, f"Напоминание: {event[0]} {event[1]} в {event[2]} скоро начнется! ⏰")
                logger.info(f"Отправлено напоминание пользователю {user_id} для мероприятия {event_id}")
            delete_reminder(reminder[0])
    except Exception as e:
        logger.error(f"Ошибка при проверке напоминаний: {e}")

async def check_non_attendance():
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        past_events = get_past_events(today)
        for event in past_events:
            event_id = event[0]
            non_attended = get_non_attended_registrations(event_id)
            for reg in non_attended:
                reg_id = reg[0]
                user_id = reg[1]
                keyboard = InlineKeyboardBuilder()
                keyboard.button(text="Медотвод ⚕️", callback_data=f"reason_med_{reg_id}")
                keyboard.button(text="Личные причины 👤", callback_data=f"reason_personal_{reg_id}")
                keyboard.button(text="Не захотел 😔", callback_data=f"reason_no_{reg_id}")
                await bot.send_message(user_id, "Вы зарегистрировались на прошедшее мероприятие, но не пришли. Укажите причину: ❓", reply_markup=keyboard.as_markup())
                logger.info(f"Отправлен опрос неявки пользователю {user_id} для reg {reg_id}")
    except Exception as e:
        logger.error(f"Ошибка при проверке неявок: {e}")

@dp.callback_query(lambda c: c.data.startswith('reason_'))
async def process_non_attendance_reason(callback_query: types.CallbackQuery):
    parts = callback_query.data.split('_')
    reason_type = parts[1]
    reg_id = int(parts[2])
    reason_map = {'med': 'медотвод', 'personal': 'личные причины', 'no': 'не захотел'}
    reason = reason_map.get(reason_type)
    if not reason:
        await callback_query.answer("Некорректная причина. ⚠️")
        return
    try:
        add_non_attendance_reason(reg_id, reason)
        await callback_query.answer("Причина записана. ✅")
        logger.info(f"Записана причина неявки для reg {reg_id}: {reason}")
    except Exception as e:
        logger.error(f"Ошибка записи причины неявки для reg {reg_id}: {e}")
        await callback_query.answer("Ошибка. Попробуйте позже. ⚠️")

async def schedule_checker():
    while True:
        schedule.run_pending()
        await asyncio.sleep(1)

async def on_startup(_):
    asyncio.create_task(schedule_checker())
    logger.info("Бот успешно запущен")

async def main():
    try:
        init_db()
        import_from_excel()
        schedule.every(10).minutes.do(lambda: asyncio.create_task(check_reminders()))  # Изменено на 10 мин
        schedule.every().day.at("00:00").do(lambda: asyncio.create_task(check_non_attendance()))  # Ежедневный опрос
        await dp.start_polling(bot, skip_updates=True, on_startup=on_startup)
    except Exception as e:
        logger.error(f"Критическая ошибка при запуске бота: {e}")
        raise

if __name__ == '__main__':
    asyncio.run(main())